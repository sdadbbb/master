import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from util.file_util import FileUtil


class LLMXlsxManager:
    """LLM 生成测试用例 XLSX 文件管理器"""

    API_HEADERS = ['序号', '需求模块', '用例名称', '描述', '优先级', '请求方法', '请求URL', '请求头', '请求数据', '变量提取', '断言规则']
    UI_HEADERS = ['序号', '需求模块', '用例名称', '描述', '优先级', '起始URL', '步骤配置']

    def __init__(self):
        self.output_dir = os.path.join(FileUtil.get_project_root(), 'reports', 'llm_cases')
        self.upload_dir = os.path.join(FileUtil.get_project_root(), 'config', 'llm', 'uploads')
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.upload_dir, exist_ok=True)

    def get_upload_dir(self):
        return self.upload_dir

    def save_cases_to_xlsx(self, generation_result, filename=None):
        """
        将生成的测试用例保存为 xlsx 文件
        :param generation_result: 包含 requirements_analysis, api_cases, ui_cases 的字典
        :return: 文件路径和文件名
        """
        if not filename:
            filename = f"test_cases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = os.path.join(self.output_dir, filename)
        wb = Workbook()

        # 需求分析 sheet
        ws_analysis = wb.active
        ws_analysis.title = '需求分析'
        ws_analysis['A1'] = '需求分析结果'
        ws_analysis['A1'].font = Font(bold=True, size=14)
        ws_analysis['A2'] = generation_result.get('requirements_analysis', '')
        ws_analysis.column_dimensions['A'].width = 100
        ws_analysis['A2'].alignment = Alignment(wrap_text=True, vertical='top')

        # 接口测试用例 sheet
        ws_api = wb.create_sheet('接口测试用例')
        self._write_header(ws_api, self.API_HEADERS)
        api_cases = generation_result.get('api_cases', [])
        for i, case in enumerate(api_cases, 1):
            request = case.get('request', {})
            ws_api.append([
                i,
                case.get('module', ''),
                case.get('name', ''),
                case.get('description', ''),
                case.get('priority', '中'),
                request.get('method', 'GET'),
                request.get('url', ''),
                json.dumps(request.get('headers', {}), ensure_ascii=False),
                json.dumps(request.get('data', {}), ensure_ascii=False),
                json.dumps(case.get('extract', {}), ensure_ascii=False),
                json.dumps(case.get('assert', {}), ensure_ascii=False)
            ])

        # UI 测试用例 sheet
        ws_ui = wb.create_sheet('UI测试用例')
        self._write_header(ws_ui, self.UI_HEADERS)
        ui_cases = generation_result.get('ui_cases', [])
        for i, case in enumerate(ui_cases, 1):
            ws_ui.append([
                i,
                case.get('module', ''),
                case.get('name', ''),
                case.get('description', ''),
                case.get('priority', '中'),
                case.get('url', ''),
                json.dumps(case.get('steps', []), ensure_ascii=False)
            ])

        wb.save(filepath)
        return filepath, filename

    def _write_header(self, ws, headers):
        """写入表头样式"""
        header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(15, len(header) * 2)

    def list_files(self):
        """列出所有生成的 xlsx 文件"""
        files = []
        for fname in os.listdir(self.output_dir):
            if fname.endswith('.xlsx'):
                fpath = os.path.join(self.output_dir, fname)
                files.append({
                    'filename': fname,
                    'size': os.path.getsize(fpath),
                    'created_at': datetime.fromtimestamp(os.path.getctime(fpath)).strftime('%Y-%m-%d %H:%M:%S')
                })
        files.sort(key=lambda x: x['created_at'], reverse=True)
        return files

    def get_filepath(self, filename):
        """获取文件完整路径（安全校验）"""
        safe_name = os.path.basename(filename)
        filepath = os.path.join(self.output_dir, safe_name)
        if os.path.exists(filepath):
            return filepath
        return None

    def read_cases_from_xlsx(self, filename):
        """从 xlsx 文件读取测试用例"""
        filepath = self.get_filepath(filename)
        if not filepath:
            return None

        wb = load_workbook(filepath, read_only=True, data_only=True)
        result = {
            'filename': filename,
            'requirements_analysis': '',
            'api_cases': [],
            'ui_cases': []
        }

        if '需求分析' in wb.sheetnames:
            ws = wb['需求分析']
            result['requirements_analysis'] = str(ws['A2'].value or '')

        if '接口测试用例' in wb.sheetnames:
            ws = wb['接口测试用例']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row or not row[2]:
                    continue
                result['api_cases'].append({
                    'module': row[1] or '',
                    'name': row[2],
                    'description': row[3] or '',
                    'priority': row[4] or '中',
                    'request': {
                        'method': row[5] or 'GET',
                        'url': row[6] or '',
                        'headers': self._parse_json_cell(row[7]),
                        'data': self._parse_json_cell(row[8])
                    },
                    'extract': self._parse_json_cell(row[9]),
                    'assert': self._parse_json_cell(row[10])
                })

        if 'UI测试用例' in wb.sheetnames:
            ws = wb['UI测试用例']
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row or not row[2]:
                    continue
                result['ui_cases'].append({
                    'module': row[1] or '',
                    'name': row[2],
                    'description': row[3] or '',
                    'priority': row[4] or '中',
                    'url': row[5] or '',
                    'steps': self._parse_json_cell(row[6])
                })

        wb.close()
        return result

    @staticmethod
    def _parse_json_cell(value):
        """解析单元格中的 JSON 字符串"""
        if not value:
            return {}
        if isinstance(value, (dict, list)):
            return value
        try:
            return json.loads(str(value))
        except (json.JSONDecodeError, TypeError):
            return {}
